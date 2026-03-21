from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from lib import normalize_multilabel, read_csv_rows, resolve_paths, write_csv_rows


INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _set_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(40, len(headers[col - 1]) + 2))


def _lock_formulas(ws) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)


def _build_main_workbook(weak_scenarios: List[Dict[str, str]], out_path: Path) -> None:
    classes = sorted(
        {
            label
            for row in weak_scenarios
            for label in normalize_multilabel(row.get("weak_tactics", ""))
            if label and not label.startswith("UNMAPPED::")
        }
    )

    wb = Workbook()

    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    inputs_headers = [
        "record_id",
        "scenario_name",
        "scenario_type",
        "weak_tactics_reference",
        "gold_final_tactics",
        "inclusion_flag",
        "notes",
    ]
    _set_header(ws_inputs, inputs_headers)
    for row in weak_scenarios:
        ws_inputs.append(
            [
                row.get("record_id", ""),
                row.get("scenario_name", ""),
                row.get("scenario_type", ""),
                row.get("weak_tactics", ""),
                "",
                "PENDING",
                "",
            ]
        )
    for r in range(2, ws_inputs.max_row + 1):
        ws_inputs.cell(r, 5).fill = INPUT_FILL
        ws_inputs.cell(r, 6).fill = INPUT_FILL
        ws_inputs.cell(r, 7).fill = INPUT_FILL

    ws_pred = wb.create_sheet("Predictions")
    pred_headers = ["record_id", "predicted_tactics", "prediction_source", "prediction_timestamp", "comments"]
    _set_header(ws_pred, pred_headers)
    for row in weak_scenarios:
        ws_pred.append([row.get("record_id", ""), "", "", "", ""])
    for r in range(2, ws_pred.max_row + 1):
        ws_pred.cell(r, 2).fill = INPUT_FILL
        ws_pred.cell(r, 3).fill = INPUT_FILL

    ws_matrix = wb.create_sheet("Label_Matrix")
    true_cols = [f"true::{c}" for c in classes]
    pred_cols = [f"pred::{c}" for c in classes]
    matrix_headers = ["record_id", "gold_final_tactics", "predicted_tactics"] + true_cols + pred_cols + ["TP", "FP", "FN", "Precision", "Recall", "F1"]
    _set_header(ws_matrix, matrix_headers)

    t_start = 4
    t_end = t_start + len(classes) - 1
    p_start = t_end + 1
    p_end = p_start + len(classes) - 1

    for i, row in enumerate(weak_scenarios, start=2):
        record_id = row.get("record_id", "")
        ws_matrix.append([
            record_id,
            f"=Inputs!E{i}",
            f"=Predictions!B{i}",
            *([0] * len(classes)),
            *([0] * len(classes)),
            "",
            "",
            "",
            "",
            "",
            "",
        ])

        for col in range(t_start, p_end + 1):
            ws_matrix.cell(i, col).fill = INPUT_FILL

        tp_col = p_end + 1
        fp_col = p_end + 2
        fn_col = p_end + 3
        prec_col = p_end + 4
        rec_col = p_end + 5
        f1_col = p_end + 6

        t_range = f"{ws_matrix.cell(i, t_start).coordinate}:{ws_matrix.cell(i, t_end).coordinate}"
        p_range = f"{ws_matrix.cell(i, p_start).coordinate}:{ws_matrix.cell(i, p_end).coordinate}"

        ws_matrix.cell(i, tp_col).value = f"=SUMPRODUCT({t_range},{p_range})"
        ws_matrix.cell(i, fp_col).value = f"=SUMPRODUCT(--({p_range}=1),--({t_range}=0))"
        ws_matrix.cell(i, fn_col).value = f"=SUMPRODUCT(--({t_range}=1),--({p_range}=0))"
        ws_matrix.cell(i, prec_col).value = f"=IF(({ws_matrix.cell(i, tp_col).coordinate}+{ws_matrix.cell(i, fp_col).coordinate})=0,0,{ws_matrix.cell(i, tp_col).coordinate}/({ws_matrix.cell(i, tp_col).coordinate}+{ws_matrix.cell(i, fp_col).coordinate}))"
        ws_matrix.cell(i, rec_col).value = f"=IF(({ws_matrix.cell(i, tp_col).coordinate}+{ws_matrix.cell(i, fn_col).coordinate})=0,0,{ws_matrix.cell(i, tp_col).coordinate}/({ws_matrix.cell(i, tp_col).coordinate}+{ws_matrix.cell(i, fn_col).coordinate}))"
        ws_matrix.cell(i, f1_col).value = f"=IF(({ws_matrix.cell(i, prec_col).coordinate}+{ws_matrix.cell(i, rec_col).coordinate})=0,0,2*{ws_matrix.cell(i, prec_col).coordinate}*{ws_matrix.cell(i, rec_col).coordinate}/({ws_matrix.cell(i, prec_col).coordinate}+{ws_matrix.cell(i, rec_col).coordinate}))"

    ws_metrics = wb.create_sheet("Metrics")
    _set_header(ws_metrics, ["Metric", "Value", "Formula/Definition"])

    last = ws_matrix.max_row
    tp_col = p_end + 1
    fp_col = p_end + 2
    fn_col = p_end + 3
    f1_col = p_end + 6

    metrics_rows = [
        ("Micro TP", f"=SUM(Label_Matrix!{ws_matrix.cell(2, tp_col).coordinate}:Label_Matrix!{ws_matrix.cell(last, tp_col).coordinate})", "sum of per-row TP"),
        ("Micro FP", f"=SUM(Label_Matrix!{ws_matrix.cell(2, fp_col).coordinate}:Label_Matrix!{ws_matrix.cell(last, fp_col).coordinate})", "sum of per-row FP"),
        ("Micro FN", f"=SUM(Label_Matrix!{ws_matrix.cell(2, fn_col).coordinate}:Label_Matrix!{ws_matrix.cell(last, fn_col).coordinate})", "sum of per-row FN"),
        ("Micro Precision", "=IF((B2+B3)=0,0,B2/(B2+B3))", "TP/(TP+FP)"),
        ("Micro Recall", "=IF((B2+B4)=0,0,B2/(B2+B4))", "TP/(TP+FN)"),
        ("Micro F1", "=IF((B5+B6)=0,0,2*B5*B6/(B5+B6))", "2PR/(P+R)"),
        ("Macro F1 (row average)", f"=AVERAGE(Label_Matrix!{ws_matrix.cell(2, f1_col).coordinate}:Label_Matrix!{ws_matrix.cell(last, f1_col).coordinate})", "average per-row F1"),
    ]
    for r in metrics_rows:
        ws_metrics.append(list(r))

    ws_cm = wb.create_sheet("Confusion_Matrix")
    _set_header(ws_cm, ["Class", "Support(True)", "Predicted_Count", "TP", "FP", "FN", "Precision", "Recall", "F1"])
    for cls in classes:
        ws_cm.append([cls, "", "", "", "", "", "", "", ""])

    for idx, _cls in enumerate(classes, start=2):
        t_col = t_start + idx - 2
        p_col = p_start + idx - 2
        ws_cm.cell(idx, 2).value = f"=SUM(Label_Matrix!{ws_matrix.cell(2, t_col).coordinate}:Label_Matrix!{ws_matrix.cell(last, t_col).coordinate})"
        ws_cm.cell(idx, 3).value = f"=SUM(Label_Matrix!{ws_matrix.cell(2, p_col).coordinate}:Label_Matrix!{ws_matrix.cell(last, p_col).coordinate})"
        ws_cm.cell(idx, 4).value = f"=SUMPRODUCT(Label_Matrix!{ws_matrix.cell(2, t_col).coordinate}:{ws_matrix.cell(last, t_col).coordinate},Label_Matrix!{ws_matrix.cell(2, p_col).coordinate}:{ws_matrix.cell(last, p_col).coordinate})"
        ws_cm.cell(idx, 5).value = f"=C{idx}-D{idx}"
        ws_cm.cell(idx, 6).value = f"=B{idx}-D{idx}"
        ws_cm.cell(idx, 7).value = f"=IF((D{idx}+E{idx})=0,0,D{idx}/(D{idx}+E{idx}))"
        ws_cm.cell(idx, 8).value = f"=IF((D{idx}+F{idx})=0,0,D{idx}/(D{idx}+F{idx}))"
        ws_cm.cell(idx, 9).value = f"=IF((G{idx}+H{idx})=0,0,2*G{idx}*H{idx}/(G{idx}+H{idx}))"

    ws_audit = wb.create_sheet("Supervisor_Audit")
    _set_header(ws_audit, ["Step", "Action", "Where", "Expected output"])
    audit_rows = [
        (1, "Fill Inputs!gold_final_tactics for validated records", "Inputs sheet", "Gold labels completed"),
        (2, "Paste model predictions in Predictions!predicted_tactics", "Predictions sheet", "Predictions completed"),
        (3, "Convert tactics to binary matrix (0/1)", "Label_Matrix true::*/pred::* columns", "Binary matrix completed"),
        (4, "Review per-row TP/FP/FN/Precision/Recall/F1", "Label_Matrix", "Per-record metrics"),
        (5, "Review global micro/macro metrics", "Metrics", "Final thesis numbers"),
        (6, "Review per-class metrics and supports", "Confusion_Matrix", "Class-level evidence"),
    ]
    for row in audit_rows:
        ws_audit.append(list(row))

    ws_dict = wb.create_sheet("Data_Dictionary")
    _set_header(ws_dict, ["Field", "Meaning", "Allowed values"])
    dictionary_rows = [
        ("gold_final_tactics", "Manually validated final labels", "semicolon-separated taxonomy tactics"),
        ("predicted_tactics", "Model output labels", "semicolon-separated taxonomy tactics"),
        ("inclusion_flag", "Whether record is part of thesis metric set", "INCLUDE/EXCLUDE/PENDING"),
        ("true::class", "Binary true indicator for class", "0 or 1"),
        ("pred::class", "Binary predicted indicator for class", "0 or 1"),
    ]
    for row in dictionary_rows:
        ws_dict.append(list(row))

    for ws in [ws_matrix, ws_metrics, ws_cm]:
        _lock_formulas(ws)
        ws.protection.sheet = True
        ws.protection.password = "supervisor"

    for ws in [ws_inputs, ws_pred, ws_audit, ws_dict]:
        ws.protection.sheet = False

    wb.save(out_path)


def _build_confusion_template(classes: List[str], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Confusion_Template"
    _set_header(ws, ["gold_label", "predicted_label", "count"])
    for gold in classes:
        for pred in classes:
            ws.append([gold, pred, ""])
            ws.cell(ws.max_row, 3).fill = INPUT_FILL
    wb.save(out_path)


def main() -> None:
    paths = resolve_paths()
    weak_scenarios = read_csv_rows(paths.data_dir / "weak_labels_scenarios.csv")

    predictions_template_path = paths.evaluation_dir / "predictions_template.csv"
    write_csv_rows(
        predictions_template_path,
        [
            {
                "record_id": row.get("record_id", ""),
                "predicted_labels": "",
                "predicted_tactics": "",
                "prediction_source": "",
                "prediction_timestamp": "",
                "comments": "",
            }
            for row in weak_scenarios
        ],
        [
            "record_id",
            "predicted_labels",
            "predicted_tactics",
            "prediction_source",
            "prediction_timestamp",
            "comments",
        ],
    )

    workbook_path = paths.evaluation_dir / "f1_recompute_template.xlsx"
    _build_main_workbook(weak_scenarios, workbook_path)

    classes = sorted(
        {
            label
            for row in weak_scenarios
            for label in normalize_multilabel(row.get("weak_tactics", ""))
            if label and not label.startswith("UNMAPPED::")
        }
    )
    confusion_template_path = paths.evaluation_dir / "confusion_matrix_template.xlsx"
    _build_confusion_template(classes, confusion_template_path)

    print(f"Wrote {predictions_template_path}")
    print(f"Wrote {workbook_path}")
    print(f"Wrote {confusion_template_path}")


if __name__ == "__main__":
    main()
